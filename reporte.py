"""
Genera el reporte final del bot: un Excel "Facturas PLCOLAB <fecha>.xlsx" con el
formato del ejemplo (Reporte facturas ...xlsx) y un zip con todos los XML.

Columnas: CLIENTE · FACTURA UT · CANTIDAD DE REMESAS · REMESAS ASOCIADAS ·
          ¿Subió al RNDC? · Novedad
"""
import io
import os
import re
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["FECHA DE EMISION", "FECHA ENVIO", "CLIENTE", "FACTURA UT",
           "CANTIDAD DE REMESAS", "REMESAS ASOCIADAS", "¿Subió al RNDC?", "Novedad"]
ANCHOS = [16.5, 16.5, 23.6, 18.7, 17.9, 30.0, 16.0, 55.0]

_AZUL = "5B9BD5"       # encabezado (tema accent5 del ejemplo)
_BLANCO = "FFFFFFFF"
_ROJO = "FFC00000"     # novedad de error
_VERDE = "FF006100"    # novedad de éxito


def _ingresoid_de(mensaje):
    """Extrae el radicado/ingresoid del mensaje de éxito ('Radicado RNDC: 12345')."""
    m = re.search(r"Radicado\s+RNDC:\s*(\d+)", mensaje or "")
    return m.group(1) if m else ""


def fecha_archivo(fecha):
    """date/str -> 'DD-MM-YYYY' (como el ejemplo 'Reporte facturas 22-05-2026')."""
    if hasattr(fecha, "strftime"):
        return fecha.strftime("%d-%m-%Y")
    s = str(fecha)
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else s


def generar_reporte_excel(facturas, fecha, out_dir, fecha_envio=None):
    """`facturas` = lista de dicts:
        {cliente, factura, remesas:[consec...], exito:bool, ingresoid:str|None, mensaje:str}
    `fecha` = fecha de EMISIÓN de las facturas (la de la consulta). `fecha_envio` = día
    en que el bot las envió al RNDC (default: hoy). Escribe
    'Facturas PLCOLAB <DD-MM-YYYY>.xlsx' en out_dir. Retorna la ruta."""
    from datetime import date as _date
    f_emision = fecha_archivo(fecha)
    f_envio = fecha_archivo(fecha_envio or _date.today())
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.title = "Hoja1"
    ws = wb.active
    ws.title = "Hoja1"

    header_fill = PatternFill("solid", fgColor=_AZUL)
    header_font = Font(bold=True, color=_BLANCO, name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFBFBFBF")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezados en la PRIMERA fila (sin espacio arriba).
    fila_h = 1
    ws.freeze_panes = "A2"   # deja fija la fila de encabezados al hacer scroll
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=fila_h, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = ANCHOS[i - 1]

    r = fila_h + 1
    for f in facturas:
        exito = bool(f.get("exito"))
        remesas = f.get("remesas", []) or []
        ingresoid = f.get("ingresoid") or _ingresoid_de(f.get("mensaje", ""))
        if exito:
            novedad = "Factura electronica grabada con éxito."
            if ingresoid:
                novedad += f" Radicado: {ingresoid}"
        else:
            novedad = f.get("mensaje", "") or "No subió."
        factura = f.get("factura", "")
        factura_val = int(factura) if str(factura).isdigit() else factura
        valores = [
            f_emision,
            f_envio,
            f.get("cliente", ""),
            factura_val,
            len(remesas),
            ", ".join(str(x) for x in remesas),
            "SI" if exito else "NO",
            novedad,
        ]
        col_novedad = len(HEADERS)  # Novedad = última columna
        for i, v in enumerate(valores, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = center
            c.border = borde
            if i == col_novedad:  # Novedad: rojo si error, verde si éxito
                c.font = Font(color=_VERDE if exito else _ROJO, bold=not exito)
        r += 1

    # ── Totales al final ──────────────────────────────────────────────────────
    total_facturas = len(facturas)
    total_remesas = sum(len(f.get("remesas", []) or []) for f in facturas)
    total_si = sum(1 for f in facturas if f.get("exito"))
    total_no = total_facturas - total_si

    r += 1  # una fila en blanco antes de los totales
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    resumen = [
        ("TOTAL DE FACTURAS:", total_facturas),
        ("TOTAL DE REMESAS:", total_remesas),
        ("TOTAL SUBIDAS (SI):", total_si),
        ("TOTAL NO SUBIDAS (NO):", total_no),
    ]
    for etiqueta, valor in resumen:
        ce = ws.cell(row=r, column=1, value=etiqueta)
        ce.font = bold
        ce.alignment = left
        cv = ws.cell(row=r, column=2, value=valor)
        cv.font = bold
        cv.alignment = left
        r += 1

    ruta = os.path.join(out_dir, f"Facturas PLCOLAB {fecha_archivo(fecha)}.xlsx")
    wb.save(ruta)
    return ruta


def guardar_datos_rg(df, fecha, out_dir):
    """Guarda el DataFrame `datos_rg` (el que alimenta la generación) como Excel,
    para auditoría — equivale al datos_rg.xlsx del flujo manual. Retorna la ruta."""
    os.makedirs(out_dir, exist_ok=True)
    ruta = os.path.join(out_dir, f"datos_rg PLCOLAB {fecha_archivo(fecha)}.xlsx")
    df.to_excel(ruta, index=False)
    return ruta


def generar_zip_xmls(xmls, fecha, out_dir):
    """`xmls` = lista de (nombre_archivo, bytes). Escribe un zip con todos. Retorna la ruta."""
    os.makedirs(out_dir, exist_ok=True)
    ruta = os.path.join(out_dir, f"Facturas PLCOLAB {fecha_archivo(fecha)}.zip")
    with zipfile.ZipFile(ruta, "w", zipfile.ZIP_DEFLATED) as zf:
        for nombre, b in xmls:
            zf.writestr(nombre, b)
    return ruta
